import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "data" / "destinations.json"
OUTPUT_PATH = ROOT / "data" / "vacation-destination-master.xlsx"

parsed = json.loads(DATA_PATH.read_text())
destinations = parsed["destinations"] if isinstance(parsed, dict) else parsed

score_keys = [
    "beach", "lake", "mountain", "woods", "desert", "city", "historic", "collegeTown",
    "sportsTown", "themePark", "island", "resort", "ski", "remote", "nightlife",
    "family", "crowds", "quiet", "luxury", "budget", "outdoorAdventure", "foodScene",
    "museums", "shows", "shopping", "kidActivities", "waterActivities", "extremeSports",
    "arts", "stargazing", "romantic", "hiddenGem", "touristy", "rvFriendly", "camping",
    "hotel", "cabin", "airbnb",
]

headers = [
    "id", "name", "region", "country", "domestic", "lat", "lon", "category",
    "environments", "activities", "lodging", "idealMinDays", "idealMaxDays",
    "sleepBudgetPerNight", "travelModes", "seasonTags", "tripStyleTags",
    *[f"score_{key}" for key in score_keys],
    "summary", "masterNotes",
]

wb = Workbook()
ws = wb.active
ws.title = "Destinations"
guide = wb.create_sheet("Tag Guide")
summary = wb.create_sheet("Summary")

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for dest in destinations:
    scores = dest.get("tagScores", {})
    row = [
        dest.get("id"),
        dest.get("name"),
        dest.get("region"),
        dest.get("country"),
        dest.get("domestic"),
        dest.get("lat"),
        dest.get("lon"),
        dest.get("category"),
        ", ".join(dest.get("environments", [])),
        ", ".join(dest.get("activities", [])),
        ", ".join(dest.get("lodging", [])),
        dest.get("idealMinDays"),
        dest.get("idealMaxDays"),
        dest.get("sleepBudgetPerNight"),
        ", ".join(dest.get("travelModes", [])),
        ", ".join(dest.get("seasonTags", [])),
        ", ".join(dest.get("tripStyleTags", [])),
        *[scores.get(key, 0) for key in score_keys],
        dest.get("summary"),
        dest.get("masterNotes"),
    ]
    ws.append(row)

last_row = ws.max_row
last_col = ws.max_column
table = Table(displayName="DestinationMaster", ref=f"A1:{get_column_letter(last_col)}{last_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(table)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

widths = {
    "A": 26, "B": 28, "C": 14, "D": 18, "E": 10, "F": 10, "G": 10, "H": 14,
    "I": 26, "J": 42, "K": 32, "L": 12, "M": 12, "N": 16, "O": 30,
    "P": 24, "Q": 32,
}
for col_idx in range(1, last_col + 1):
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = widths.get(letter, 12 if col_idx < 18 else 13)

for row in ws.iter_rows(min_row=2, max_row=last_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

score_start = headers.index("score_beach") + 1
score_end = headers.index("score_airbnb") + 1
score_range = f"{get_column_letter(score_start)}2:{get_column_letter(score_end)}{last_row}"
ws.conditional_formatting.add(score_range, ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=3, mid_color="FFEB84", end_type="num", end_value=5, end_color="63BE7B"))

guide_rows = [
    ["Score", "Meaning"],
    [0, "Not a meaningful fit for this destination"],
    [1, "Weak fit"],
    [2, "Possible but not a defining trait"],
    [3, "Moderate fit"],
    [4, "Strong fit"],
    [5, "Defining trait / very strong fit"],
    [],
    ["Column Group", "How to use it"],
    ["Basic fields", "Name, region, country, coordinates, trip length, sleep-only lodging target"],
    ["Array fields", "Comma-separated editable tags for environments, activities, lodging, travel modes, seasons, and trip styles"],
    ["score_* fields", "0-5 trait scores used by the AI and local pre-ranking context"],
    ["Distance", "Not stored here. The app calculates distance live from each user's ZIP code."],
]
for row in guide_rows:
    guide.append(row)
for cell in guide[1]:
    cell.fill = header_fill
    cell.font = header_font
guide.column_dimensions["A"].width = 22
guide.column_dimensions["B"].width = 95
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

summary.append(["Metric", "Value"])
summary.append(["Total destinations", len(destinations)])
summary.append(["United States destinations", sum(1 for d in destinations if d.get("domestic"))])
summary.append(["Worldwide destinations", sum(1 for d in destinations if not d.get("domestic"))])
summary.append(["Tag version", parsed.get("meta", {}).get("tagVersion", 2) if isinstance(parsed, dict) else 2])
summary.append(["Score scale", "0-5"])
summary.append(["Generated from", "data/destinations.json"])
for cell in summary[1]:
    cell.fill = header_fill
    cell.font = header_font
summary.column_dimensions["A"].width = 28
summary.column_dimensions["B"].width = 42
for row in summary.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

wb.save(OUTPUT_PATH)
print(f"Wrote {OUTPUT_PATH}")
